# app.py
# - 설정 추가 (문제/보기 순서 선택)
# - 보기/문제 셔플 후 번호 정렬 및 정답 매핑 보존
from flask import Flask, render_template, request, redirect, url_for, flash, send_file, session
from openpyxl import Workbook, load_workbook
import os
import tempfile
from werkzeug.utils import secure_filename
from werkzeug.middleware.proxy_fix import ProxyFix
from random import shuffle

app = Flask(__name__)
app.secret_key = 'your-secret-key-here'
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024
app.wsgi_app = ProxyFix(app.wsgi_app, x_prefix=1)

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() == 'xlsx'

def read_quiz_file(filepath):
    try:
        workbook = load_workbook(filepath)
        sheet = workbook.active
        questions = []
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 1):
            if row[0] is None:
                break
            question = {
                'id': row_num,
                'question': str(row[0]),
                'choice1': str(row[1]),
                'choice2': str(row[2]),
                'choice3': str(row[3]),
                'choice4': str(row[4]),
                'answer': int(row[5]),
                'explanation': str(row[6]) if row[6] else ''
            }
            questions.append(question)
        return questions
    except Exception as e:
        print(f"파일 읽기 오류: {e}")
        return None

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/download_template')
def download_template():
    workbook = Workbook()
    sheet = workbook.active
    headers = ['문제', '선택지1', '선택지2', '선택지3', '선택지4', '정답', '해설']
    for col, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col, value=header)
    sample_data = [
        ['Python에서 리스트를 정의하는 기호는?', '()', '[]', '{}', '""', 2, '대괄호 []를 사용합니다.'],
        ['2 + 3 * 4의 결과는?', '14', '20', '11', '9', 1, '곱셈이 덧셈보다 우선순위가 높습니다. 2 + (3*4) = 14'],
        ['HTML의 제목 태그는?', '<title>', '<h1>', '<head>', '<header>', 2, '<h1>은 가장 큰 제목 태그입니다.']
    ]
    for row_num, data in enumerate(sample_data, 2):
        for col, value in enumerate(data, 1):
            sheet.cell(row=row_num, column=col, value=value)
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        workbook.save(tmp.name)
        return send_file(tmp.name, as_attachment=True, download_name='quiz_template.xlsx')

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        flash('파일을 선택해주세요.')
        return redirect(url_for('index'))
    file = request.files['file']
    if file.filename == '':
        flash('파일을 선택해주세요.')
        return redirect(url_for('index'))
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        questions = read_quiz_file(filepath)
        if questions:
            session['questions'] = questions
            session['filename'] = filename
            flash(f'파일이 성공적으로 업로드되었습니다! ({len(questions)}문제)')
            return redirect(url_for('index'))
        else:
            flash('파일 형식이 올바르지 않습니다. 템플릿을 확인해주세요.')
            return redirect(url_for('index'))
    else:
        flash('Excel 파일만 업로드 가능합니다.')
        return redirect(url_for('index'))

def prepare_questions(mode):
    questions = session.get('questions', [])
    settings = session.get('settings', {}).get(mode, {})
    question_order = settings.get('question_order', 'original')
    choice_order = settings.get('choice_order', 'original')
    limit = settings.get('limit', 0)

    processed = questions.copy()
    if question_order == 'shuffle':
        shuffle(processed)
    if limit and 1 <= limit < len(processed):
        processed = processed[:limit]

    for display_index, q in enumerate(processed, 1):
        q['display_id'] = display_index
        choices = [q['choice1'], q['choice2'], q['choice3'], q['choice4']]
        indexed = list(enumerate(choices, start=1))
        if choice_order == 'shuffle':
            shuffle(indexed)
        q['shuffled_choices'] = [
            {'display_index': i + 1, 'text': text, 'original_index': orig_idx}
            for i, (orig_idx, text) in enumerate(indexed)
        ]
    return processed

@app.route('/study')
def study():
    if not session.get('questions'):
        flash('먼저 문제 파일을 업로드해주세요.')
        return redirect(url_for('index'))
    questions = prepare_questions('study')
    return render_template('study.html', questions=questions)

@app.route('/quiz')
def quiz():
    if not session.get('questions'):
        flash('먼저 문제 파일을 업로드해주세요.')
        return redirect(url_for('index'))
    questions = prepare_questions('quiz')
    session['quiz_data'] = {q['display_id']: q for q in questions}  # 채점용 백업
    return render_template('quiz.html', questions=questions)

@app.route('/submit_quiz', methods=['POST'])
def submit_quiz():
    quiz_data = session.get('quiz_data')
    if not quiz_data:
        flash('시험 세션이 만료되었습니다. 다시 시작해주세요.')
        return redirect(url_for('index'))

    results = []
    correct_count = 0

    for display_id, question in quiz_data.items():
        user_display_answer = request.form.get(f'question_{display_id}')
        is_correct = False
        user_real_answer = None

        if user_display_answer:
            user_display_answer = int(user_display_answer)
            display_to_original = {
                item['display_index']: item['original_index']
                for item in question['shuffled_choices']
            }
            user_real_answer = display_to_original.get(user_display_answer)
            is_correct = user_real_answer == question['answer']
            if is_correct:
                correct_count += 1

        results.append({
            'question': question,
            'user_answer': user_display_answer,
            'real_answer': user_real_answer,
            'is_correct': is_correct
        })

    score_percentage = round((correct_count / len(quiz_data)) * 100, 1)
    return render_template('result.html', results=results, correct_count=correct_count,
                           total_questions=len(quiz_data), score_percentage=score_percentage)

@app.route('/settings', methods=['GET', 'POST'])
def settings():
    if request.method == 'POST':
        session['settings'] = {
            'study': {
                'question_order': request.form.get('study_question_order', 'original'),
                'choice_order': request.form.get('study_choice_order', 'original'),
                'limit': int(request.form.get('study_limit', 0) or 0)
            },
            'quiz': {
                'question_order': request.form.get('quiz_question_order', 'original'),
                'choice_order': request.form.get('quiz_choice_order', 'original'),
                'limit': int(request.form.get('quiz_limit', 0) or 0)
            }
        }
        flash('설정이 저장되었습니다.')
        return redirect(url_for('settings'))

    settings = session.get('settings', {})
    return render_template('settings.html', settings=settings)

if __name__ == '__main__':
    app.run(debug=True)
