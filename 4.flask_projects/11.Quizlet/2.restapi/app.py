from flask import Flask, request, jsonify, send_from_directory, session, send_file
from flask_cors import CORS
from openpyxl import Workbook, load_workbook
import os
import tempfile
import uuid
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.secret_key = 'your-secret-key-here'
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# CORS 설정
CORS(app)

# 업로드 폴더 생성
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() == 'xlsx'

def read_quiz_file(filepath):
    """Excel 파일에서 퀴즈 데이터 읽기"""
    try:
        workbook = load_workbook(filepath)
        sheet = workbook.active
        questions = []
        
        # 헤더 건너뛰고 데이터 읽기 (2번째 행부터)
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 1):
            if row[0] is None:  # 빈 행이면 중단
                break
                
            question = {
                'id': row_num,
                'question': str(row[0]),
                'choice1': str(row[1]),
                'choice2': str(row[2]),
                'choice3': str(row[3]),
                'choice4': str(row[4]),
                'answer': int(row[5]),  # 1, 2, 3, 4 중 하나
                'explanation': str(row[6]) if row[6] else ''
            }
            questions.append(question)
        
        return questions
    except Exception as e:
        return None

# 정적 파일 서빙 (메인 페이지)
@app.route('/')
def index():
    return send_from_directory('static', 'index.html')

@app.route('/static/<path:filename>')
def static_files(filename):
    return send_from_directory('static', filename)

# REST API 엔드포인트들
@app.route('/api/download-template', methods=['GET'])
def api_download_template():
    """샘플 템플릿 다운로드"""
    try:
        # 새 워크북 생성
        workbook = Workbook()
        sheet = workbook.active
        
        # 헤더 설정
        headers = ['문제', '선택지1', '선택지2', '선택지3', '선택지4', '정답', '해설']
        for col, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col, value=header)
        
        # 샘플 데이터 추가
        sample_data = [
            ['Python에서 리스트를 정의하는 기호는?', '()', '[]', '{}', '""', 2, '대괄호 []를 사용합니다.'],
            ['2 + 3 * 4의 결과는?', '14', '20', '11', '9', 1, '곱셈이 덧셈보다 우선순위가 높습니다. 2 + (3*4) = 14'],
            ['HTML의 제목 태그는?', '<title>', '<h1>', '<head>', '<header>', 2, '<h1>은 가장 큰 제목 태그입니다.']
        ]
        
        for row_num, data in enumerate(sample_data, 2):
            for col, value in enumerate(data, 1):
                sheet.cell(row=row_num, column=col, value=value)
        
        # 임시 파일로 저장
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        workbook.save(temp_file.name)
        
        return send_file(temp_file.name, as_attachment=True, download_name='quiz_template.xlsx')
        
    except Exception as e:
        return jsonify({'error': f'템플릿 생성 실패: {str(e)}'}), 500

@app.route('/api/upload', methods=['POST'])
def api_upload_file():
    """파일 업로드 및 처리"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': '파일을 선택해주세요.'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': '파일을 선택해주세요.'}), 400
        
        if not allowed_file(file.filename):
            return jsonify({'error': 'Excel 파일만 업로드 가능합니다.'}), 400
        
        # 파일 저장
        filename = secure_filename(file.filename)
        unique_filename = f"{uuid.uuid4()}_{filename}"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], unique_filename)
        file.save(filepath)
        
        # 파일 읽기 시도
        questions = read_quiz_file(filepath)
        if questions:
            session['questions'] = questions
            session['filename'] = filename
            return jsonify({
                'success': True,
                'message': f'파일이 성공적으로 업로드되었습니다! ({len(questions)}문제)',
                'question_count': len(questions)
            })
        else:
            return jsonify({'error': '파일 형식이 올바르지 않습니다. 템플릿을 확인해주세요.'}), 400
            
    except Exception as e:
        return jsonify({'error': f'파일 업로드 실패: {str(e)}'}), 500

@app.route('/api/questions', methods=['GET'])
def api_get_questions():
    """문제 목록 조회 (공부 모드용)"""
    questions = session.get('questions')
    if not questions:
        return jsonify({'error': '먼저 문제 파일을 업로드해주세요.'}), 400
    
    return jsonify({'questions': questions})

@app.route('/api/quiz', methods=['GET'])
def api_get_quiz():
    """퀴즈 문제 조회 (시험 모드용 - 정답 제외)"""
    questions = session.get('questions')
    if not questions:
        return jsonify({'error': '먼저 문제 파일을 업로드해주세요.'}), 400
    
    # 정답과 해설 제거
    quiz_questions = []
    for q in questions:
        quiz_questions.append({
            'id': q['id'],
            'question': q['question'],
            'choice1': q['choice1'],
            'choice2': q['choice2'],
            'choice3': q['choice3'],
            'choice4': q['choice4']
        })
    
    return jsonify({'questions': quiz_questions})

@app.route('/api/submit-quiz', methods=['POST'])
def api_submit_quiz():
    """퀴즈 제출 및 채점"""
    try:
        questions = session.get('questions')
        if not questions:
            return jsonify({'error': '먼저 문제 파일을 업로드해주세요.'}), 400
        
        answers = request.json.get('answers', {})
        
        results = []
        correct_count = 0
        
        for question in questions:
            user_answer = answers.get(str(question['id']))
            is_correct = False
            
            if user_answer:
                user_answer = int(user_answer)
                is_correct = user_answer == question['answer']
                if is_correct:
                    correct_count += 1
            
            results.append({
                'question': question,
                'user_answer': user_answer,
                'is_correct': is_correct
            })
        
        score_percentage = round((correct_count / len(questions)) * 100, 1)
        
        return jsonify({
            'results': results,
            'correct_count': correct_count,
            'total_questions': len(questions),
            'score_percentage': score_percentage
        })
        
    except Exception as e:
        return jsonify({'error': f'채점 실패: {str(e)}'}), 500

@app.route('/api/status', methods=['GET'])
def api_status():
    """현재 상태 확인"""
    questions = session.get('questions')
    return jsonify({
        'has_questions': bool(questions),
        'question_count': len(questions) if questions else 0,
        'filename': session.get('filename', '')
    })

if __name__ == '__main__':
    app.run(debug=True)
