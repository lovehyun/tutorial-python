from flask import Flask, render_template, request, redirect, url_for, flash, send_file, session
from openpyxl import Workbook, load_workbook
import os
import tempfile
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.secret_key = 'your-secret-key-here'
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# 업로드 폴더 생성
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() == 'xlsx'

def read_quiz_file(filepath):
    """Excel 파일에서 퀴즈 데이터 읽기"""
    try:
        workbook = load_workbook(filepath)
        sheet = workbook.active
        questions = []
        
        # 헤더 건너뛰고 데이터 읽기 (2번째 행부터)
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 1):
            if row[0] is None:  # 빈 행이면 중단
                break
                
            question = {
                'id': row_num,
                'question': str(row[0]),
                'choice1': str(row[1]),
                'choice2': str(row[2]),
                'choice3': str(row[3]),
                'choice4': str(row[4]),
                'answer': int(row[5]),  # 1, 2, 3, 4 중 하나
                'explanation': str(row[6]) if row[6] else ''
            }
            questions.append(question)
        
        return questions
    except Exception as e:
        print(f"파일 읽기 오류: {e}")
        return None

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/download_template')
def download_template():
    """샘플 템플릿 다운로드"""
    # 새 워크북 생성
    workbook = Workbook()
    sheet = workbook.active
    
    # 헤더 설정
    headers = ['문제', '선택지1', '선택지2', '선택지3', '선택지4', '정답', '해설']
    for col, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col, value=header)
    
    # 샘플 데이터 추가
    sample_data = [
        ['Python에서 리스트를 정의하는 기호는?', '()', '[]', '{}', '""', 2, '대괄호 []를 사용합니다.'],
        ['2 + 3 * 4의 결과는?', '14', '20', '11', '9', 1, '곱셈이 덧셈보다 우선순위가 높습니다. 2 + (3*4) = 14'],
        ['HTML의 제목 태그는?', '<title>', '<h1>', '<head>', '<header>', 2, '<h1>은 가장 큰 제목 태그입니다.']
    ]
    
    for row_num, data in enumerate(sample_data, 2):
        for col, value in enumerate(data, 1):
            sheet.cell(row=row_num, column=col, value=value)
    
    # 임시 파일로 저장
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        workbook.save(tmp.name)
        return send_file(tmp.name, as_attachment=True, download_name='quiz_template.xlsx')

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        flash('파일을 선택해주세요.')
        return redirect(url_for('index'))
    
    file = request.files['file']
    if file.filename == '':
        flash('파일을 선택해주세요.')
        return redirect(url_for('index'))
    
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        # 파일 읽기 시도
        questions = read_quiz_file(filepath)
        if questions:
            session['questions'] = questions
            session['filename'] = filename
            flash(f'파일이 성공적으로 업로드되었습니다! ({len(questions)}문제)')
            return redirect(url_for('index'))
        else:
            flash('파일 형식이 올바르지 않습니다. 템플릿을 확인해주세요.')
            return redirect(url_for('index'))
    else:
        flash('Excel 파일만 업로드 가능합니다.')
        return redirect(url_for('index'))

@app.route('/study')
def study():
    """공부 모드 - 문제와 정답을 모두 보여줌"""
    questions = session.get('questions')
    if not questions:
        flash('먼저 문제 파일을 업로드해주세요.')
        return redirect(url_for('index'))
    
    return render_template('study.html', questions=questions)

@app.route('/quiz')
def quiz():
    """시험 모드 - 문제만 보여주고 답 숨김"""
    questions = session.get('questions')
    if not questions:
        flash('먼저 문제 파일을 업로드해주세요.')
        return redirect(url_for('index'))
    
    return render_template('quiz.html', questions=questions)

@app.route('/submit_quiz', methods=['POST'])
def submit_quiz():
    """시험 제출 및 채점"""
    questions = session.get('questions')
    if not questions:
        flash('먼저 문제 파일을 업로드해주세요.')
        return redirect(url_for('index'))
    
    results = []
    correct_count = 0
    
    for question in questions:
        user_answer = request.form.get(f'question_{question["id"]}')
        is_correct = False
        
        if user_answer:
            user_answer = int(user_answer)
            is_correct = user_answer == question['answer']
            if is_correct:
                correct_count += 1
        
        results.append({
            'question': question,
            'user_answer': int(user_answer) if user_answer else None,
            'is_correct': is_correct
        })
    
    score_percentage = round((correct_count / len(questions)) * 100, 1)
    
    return render_template('result.html', 
                         results=results, 
                         correct_count=correct_count,
                         total_questions=len(questions),
                         score_percentage=score_percentage)

if __name__ == '__main__':
    app.run(debug=True)
