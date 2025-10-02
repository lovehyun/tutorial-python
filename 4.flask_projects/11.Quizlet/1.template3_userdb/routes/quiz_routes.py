# routes/quiz_routes.py
# 퀴즈 관련 라우트 (파일 업로드, 문제 풀이, 설정 등)

from flask import Blueprint, render_template, request, redirect, url_for, flash, session, send_file
from flask_login import login_required, current_user
from werkzeug.utils import secure_filename
from openpyxl import Workbook, load_workbook
import os
import tempfile
from random import shuffle
from database import get_db_connection, create_user_upload_folder, create_quiz_session, get_quiz_session, deactivate_quiz_session

quiz_bp = Blueprint('quiz', __name__)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() == 'xlsx'

def read_quiz_file(filepath):
    """Excel 파일에서 퀴즈 데이터 읽기"""
    try:
        workbook = load_workbook(filepath)
        sheet = workbook.active
        questions = []
        
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 1):
            if row[0] is None:  # 문제가 없으면 중단
                break
            
            question = {
                'id': row_num,
                'question': str(row[0]),
                'choice1': str(row[1]),
                'choice2': str(row[2]),
                'choice3': str(row[3]),
                'choice4': str(row[4]),
                'answer': int(row[5]),
                'explanation': str(row[6]) if row[6] else ''
            }
            questions.append(question)
        
        return questions
    except Exception as e:
        print(f"파일 읽기 오류: {e}")
        return None

def get_user_settings(user_id, mode):
    """사용자 설정 가져오기"""
    conn = get_db_connection()
    settings = conn.execute(
        'SELECT * FROM user_settings WHERE user_id = ? AND mode = ?',
        (user_id, mode)
    ).fetchone()
    conn.close()
    
    if settings:
        return {
            'question_order': settings['question_order'],
            'choice_order': settings['choice_order'],
            'question_limit': settings['question_limit']
        }
    return {
        'question_order': 'original',
        'choice_order': 'original', 
        'question_limit': 0
    }

def prepare_questions(questions, mode, user_id):
    """설정에 따라 문제 준비하기"""
    settings = get_user_settings(user_id, mode)
    
    processed = questions.copy()
    
    # 문제 순서 처리
    if settings['question_order'] == 'shuffle':
        shuffle(processed)
    
    # 문제 수 제한
    limit = settings['question_limit']
    if limit and 1 <= limit < len(processed):
        processed = processed[:limit]
    
    # 각 문제에 display_id 부여 및 보기 섞기
    for display_index, q in enumerate(processed, 1):
        q['display_id'] = display_index
        choices = [q['choice1'], q['choice2'], q['choice3'], q['choice4']]
        indexed = list(enumerate(choices, start=1))
        
        if settings['choice_order'] == 'shuffle':
            shuffle(indexed)
        
        q['shuffled_choices'] = [
            {'display_index': i + 1, 'text': text, 'original_index': orig_idx}
            for i, (orig_idx, text) in enumerate(indexed)
        ]
    
    return processed

@quiz_bp.route('/dashboard')
@login_required
def dashboard():
    """메인 대시보드"""
    conn = get_db_connection()
    
    # 사용자의 퀴즈 파일 목록
    files = conn.execute(
        'SELECT * FROM quiz_files WHERE user_id = ? ORDER BY created_at DESC',
        (current_user.id,)
    ).fetchall()
    
    # 최근 시험 결과 5개
    recent_results = conn.execute('''
        SELECT qr.*, qf.original_filename 
        FROM quiz_results qr 
        JOIN quiz_files qf ON qr.quiz_file_id = qf.id 
        WHERE qr.user_id = ? 
        ORDER BY qr.created_at DESC 
        LIMIT 5
    ''', (current_user.id,)).fetchall()
    
    conn.close()
    
    return render_template('index.html', files=files, recent_results=recent_results)

@quiz_bp.route('/download_template')
def download_template():
    """퀴즈 템플릿 다운로드"""
    workbook = Workbook()
    sheet = workbook.active
    
    # 헤더 설정
    headers = ['문제', '선택지1', '선택지2', '선택지3', '선택지4', '정답', '해설']
    for col, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col, value=header)
    
    # 샘플 데이터
    sample_data = [
        ['Python에서 리스트를 정의하는 기호는?', '()', '[]', '{}', '""', 2, '대괄호 []를 사용합니다.'],
        ['2 + 3 * 4의 결과는?', '14', '20', '11', '9', 1, '곱셈이 덧셈보다 우선순위가 높습니다. 2 + (3*4) = 14'],
        ['HTML의 제목 태그는?', '<title>', '<h1>', '<head>', '<header>', 2, '<h1>은 가장 큰 제목 태그입니다.']
    ]
    
    for row_num, data in enumerate(sample_data, 2):
        for col, value in enumerate(data, 1):
            sheet.cell(row=row_num, column=col, value=value)
    
    # 임시 파일로 저장 후 전송
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        workbook.save(tmp.name)
        return send_file(tmp.name, as_attachment=True, download_name='quiz_template.xlsx')

@quiz_bp.route('/upload', methods=['POST'])
@login_required
def upload_file():
    """퀴즈 파일 업로드"""
    if 'file' not in request.files:
        flash('파일을 선택해주세요.', 'error')
        return redirect(url_for('quiz.dashboard'))
    
    file = request.files['file']
    if file.filename == '':
        flash('파일을 선택해주세요.', 'error')
        return redirect(url_for('quiz.dashboard'))
    
    if file and allowed_file(file.filename):
        # 사용자 폴더 확인/생성
        user_folder = create_user_upload_folder(current_user.id)
        
        # 안전한 파일명 생성
        original_filename = file.filename
        secure_name = secure_filename(original_filename)
        filepath = os.path.join(user_folder, secure_name)
        
        # 동일한 파일명 있으면 번호 추가
        counter = 1
        base_name, ext = os.path.splitext(secure_name)
        while os.path.exists(filepath):
            secure_name = f"{base_name}_{counter}{ext}"
            filepath = os.path.join(user_folder, secure_name)
            counter += 1
        
        file.save(filepath)
        
        # 파일 내용 검증
        questions = read_quiz_file(filepath)
        if questions:
            # 데이터베이스에 파일 정보 저장
            conn = get_db_connection()
            conn.execute('''
                INSERT INTO quiz_files (user_id, filename, original_filename, file_path, question_count)
                VALUES (?, ?, ?, ?, ?)
            ''', (current_user.id, secure_name, original_filename, filepath, len(questions)))
            conn.commit()
            conn.close()
            
            flash(f'파일이 성공적으로 업로드되었습니다! ({len(questions)}개 문제)', 'success')
        else:
            # 파일 삭제
            os.remove(filepath)
            flash('파일 형식이 올바르지 않습니다. 템플릿을 확인해주세요.', 'error')
    else:
        flash('Excel 파일(.xlsx)만 업로드 가능합니다.', 'error')
    
    return redirect(url_for('quiz.dashboard'))

@quiz_bp.route('/delete_file/<int:file_id>', methods=['POST'])
@login_required
def delete_file(file_id):
    """퀴즈 파일 삭제"""
    conn = get_db_connection()
    
    # 파일 소유권 확인
    file_info = conn.execute(
        'SELECT * FROM quiz_files WHERE id = ? AND user_id = ?',
        (file_id, current_user.id)
    ).fetchone()
    
    if not file_info:
        flash('파일을 찾을 수 없거나 삭제 권한이 없습니다.', 'error')
        conn.close()
        return redirect(url_for('quiz.dashboard'))
    
    try:
        # 실제 파일 삭제
        if os.path.exists(file_info['file_path']):
            os.remove(file_info['file_path'])
        
        # 데이터베이스에서 삭제 (관련 결과도 CASCADE로 삭제됨)
        conn.execute('DELETE FROM quiz_files WHERE id = ?', (file_id,))
        conn.commit()
        
        flash(f"'{file_info['original_filename']}' 파일이 삭제되었습니다.", 'success')
    except Exception as e:
        flash(f'파일 삭제 중 오류가 발생했습니다: {str(e)}', 'error')
    
    conn.close()
    return redirect(url_for('quiz.dashboard'))

@quiz_bp.route('/study/<int:file_id>')
@login_required
def study(file_id):
    """공부 모드"""
    conn = get_db_connection()
    file_info = conn.execute(
        'SELECT * FROM quiz_files WHERE id = ? AND user_id = ?',
        (file_id, current_user.id)
    ).fetchone()
    conn.close()
    
    if not file_info:
        flash('파일을 찾을 수 없습니다.', 'error')
        return redirect(url_for('quiz.dashboard'))
    
    questions = read_quiz_file(file_info['file_path'])
    if not questions:
        flash('파일을 읽을 수 없습니다.', 'error')
        return redirect(url_for('quiz.dashboard'))
    
    prepared_questions = prepare_questions(questions, 'study', current_user.id)
    return render_template('quiz/study.html', questions=prepared_questions, file_info=file_info)

@quiz_bp.route('/exam/<int:file_id>')
@login_required
def quiz_start(file_id):
    """시험 모드 시작"""
    conn = get_db_connection()
    file_info = conn.execute(
        'SELECT * FROM quiz_files WHERE id = ? AND user_id = ?',
        (file_id, current_user.id)
    ).fetchone()
    conn.close()
    
    if not file_info:
        flash('파일을 찾을 수 없습니다.', 'error')
        return redirect(url_for('quiz.dashboard'))
    
    questions = read_quiz_file(file_info['file_path'])
    if not questions:
        flash('파일을 읽을 수 없습니다.', 'error')
        return redirect(url_for('quiz.dashboard'))
    
    prepared_questions = prepare_questions(questions, 'quiz', current_user.id)
    settings = get_user_settings(current_user.id, 'quiz')
    
    # 데이터베이스에 퀴즈 세션 생성
    session_token = create_quiz_session(
        current_user.id, 
        file_id, 
        {q['display_id']: q for q in prepared_questions},
        settings
    )
    
    return render_template('quiz/quiz.html', 
                         questions=prepared_questions, 
                         file_info=file_info,
                         session_token=session_token)

@quiz_bp.route('/settings', methods=['GET', 'POST'])
@login_required
def settings():
    """설정 페이지"""
    if request.method == 'POST':
        conn = get_db_connection()
        
        # 공부 모드 설정 업데이트
        conn.execute('''
            UPDATE user_settings 
            SET question_order = ?, choice_order = ?, question_limit = ?
            WHERE user_id = ? AND mode = 'study'
        ''', (
            request.form.get('study_question_order', 'original'),
            request.form.get('study_choice_order', 'original'),
            int(request.form.get('study_limit', 0) or 0),
            current_user.id
        ))
        
        # 시험 모드 설정 업데이트
        conn.execute('''
            UPDATE user_settings 
            SET question_order = ?, choice_order = ?, question_limit = ?
            WHERE user_id = ? AND mode = 'quiz'
        ''', (
            request.form.get('quiz_question_order', 'original'),
            request.form.get('quiz_choice_order', 'original'),
            int(request.form.get('quiz_limit', 0) or 0),
            current_user.id
        ))
        
        conn.commit()
        conn.close()
        
        flash('설정이 저장되었습니다.', 'success')
        return redirect(url_for('quiz.settings'))
    
    # 현재 설정 가져오기
    conn = get_db_connection()
    settings = {}
    for mode in ['study', 'quiz']:
        setting = conn.execute(
            'SELECT * FROM user_settings WHERE user_id = ? AND mode = ?',
            (current_user.id, mode)
        ).fetchone()
        if setting:
            settings[mode] = {
                'question_order': setting['question_order'],
                'choice_order': setting['choice_order'],
                'question_limit': setting['question_limit']
            }
    conn.close()
    
    return render_template('quiz/settings.html', settings=settings)
