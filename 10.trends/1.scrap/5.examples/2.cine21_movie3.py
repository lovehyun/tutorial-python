# 2. 엑셀파일 저장

from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
import pandas as pd

import requests
from io import BytesIO
from PIL import Image
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.worksheet.hyperlink import Hyperlink


# 크롬 드라이버 경로 설정 및 headless 모드 활성화
options = webdriver.ChromeOptions()
options.add_argument('--headless')

# 크롬 드라이버 생성
driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), options=options)

# 웹 페이지에 접속
base_url = 'http://www.cine21.com/'
ranking_url = base_url + 'rank/boxoffice/domestic'
driver.get(ranking_url)

# 페이지 소스코드 가져오기
page_source = driver.page_source

# BeautifulSoup을 사용하여 페이지 소스코드 파싱
bs = BeautifulSoup(page_source, 'html.parser')

# boxoffice_list_content ID를 가진 div 찾기
boxoffice_list_content = bs.find('div', id='boxoffice_list_content')

# boxoffice_list_content 내의 모든 boxoffice_li 클래스를 가진 li 요소들 찾기
boxoffice_li_list = boxoffice_list_content.find_all('li', class_='boxoffice_li')

# 웹 드라이버 종료
driver.quit()

# 데이터 프레임 생성
data = []

# 각 boxoffice_li에 대해 mov_name, people_num 클래스를 가진 div의 텍스트 출력
for rank, boxoffice_li in enumerate(boxoffice_li_list, start=1):
    mov_name_div = boxoffice_li.find('div', class_='mov_name')
    people_num_div = boxoffice_li.find('div', class_='people_num')
    mov_name = mov_name_div.get_text(strip=True) if mov_name_div else ''
    people_num = people_num_div.get_text(strip=True).replace('관객수|', '') if people_num_div else ''

    mov_link = base_url + boxoffice_li.find('a')['href']
    img_src = boxoffice_li.find('img')['src']

    print(f"순위: {rank}, 영화 제목: {mov_name}, 관객 수: {people_num}, 웹사이트 정보: {mov_link}, 포스터: {img_src}")

    data.append({'순위': rank, '영화 제목': mov_name, '관객 수': people_num, '웹사이트 정보': mov_link, '포스터': img_src})

df = pd.DataFrame(data)

# 엑셀 파일로 저장
df.to_excel('boxoffice_rankings.xlsx', index=False)


# 이미지 삽입을 위해 직접 엑셀 워크북 생성
wb = Workbook()
ws = wb.active

# 헤더 작성
ws.append(['순위', '포스터', '영화 제목', '관객 수', '웹사이트 정보'])

# 데이터 입력
for row_num, item in enumerate(data, start=2):
    ws.cell(row=row_num, column=1, value=item['순위'])
    ws.cell(row=row_num, column=3, value=item['영화 제목'])
    # 영화 제목 셀의 폭 조정
    ws.column_dimensions['C'].width = 30

    ws.cell(row=row_num, column=4, value=item['관객 수'])
    ws.cell(row=row_num, column=5, value=item['웹사이트 정보'])
    # 하이퍼링크 추가
    ws.cell(row=row_num, column=5).hyperlink = Hyperlink(item['웹사이트 정보'])

    # 이미지 다운로드 및 삽입
    response = requests.get(item['포스터'])
    img = Image.open(BytesIO(response.content))
    img.thumbnail((100, 100))  # 이미지 크기 조정
    img_path = f"image{row_num}.jpg"
    img.save(img_path)
    img_obj = ExcelImage(img_path)
    ws.add_image(img_obj, f'B{row_num}')  # 이미지 삽입 위치

    # 각 행의 높이 설정
    ws.row_dimensions[row_num].height = 100

# 엑셀 파일 저장
wb.save('boxoffice_rankings_with_images.xlsx')
